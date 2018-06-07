#coding:utf-8
# @Time    : 2018/5/17
# @Author  : xiaopeng
# @Site    : boxueshuyuan
# @File    : test4.py
# @Software: PyCharm
import pymongo
import openpyxl
import datetime
from collections import defaultdict
from openpyxl import styles
import traceback

client = pymongo.MongoClient(host='mongodb://root:miaoji1109-=@10.19.2.103:27017/')
package_id_list = [6]
package_id_count_list = {}
today_date = datetime.datetime.strftime(datetime.datetime.now(), '%Y%m%d')

def error_code_by_source_by_rank():
    source_list = ['hotelsListHotel', 'elongListHotel', 'ctripListHotel', 'expediaListHotel', 'bookingListHotel', 'agodaListHotel']
    res = defaultdict(dict)
    for package_id in package_id_list:
        per_count = 0
        for source in source_list:
            table_name = 'DateTask_Hotel_{}_{}'.format(source, today_date)
            client['RoutineDateTask'][table_name].update({'package_id':package_id}, {'$set':{'used_times':0}}, multi=True)


    #         count =  client['RoutineDateTask'][table_name].find({'package_id': package_id, 'feedback_times':{'$gt':0}}).count()
    #         per_count += count
    #     package_id_count_list[package_id] = per_count
    # for source in source_list:
    #     res = get_dict_util(res, source)
    # # sorted(res.keys())
    # to_csv(res)
    # print()

def get_dict_util(res, source):
    try:
        table_name = 'DateTask_Hotel_{}_{}'.format(source, today_date)
        col_1 = client['RoutineDateTask'][table_name]
        for package_id in package_id_list:
            res[package_id][source] = defaultdict(dict)
            per_total_count = col_1.find({'package_id': package_id}).count()
            cursor_1 = col_1.aggregate(
                [{'$match': {'package_id': package_id}}, {'$group': {'_id': '$error_code', 'count': {'$sum': 1}}}])

            for i in cursor_1:
                if i['_id'] is not None:
                    res[package_id][source][i['_id']] = i['count']
            for _id, count in res[package_id][source].items():
                if not per_total_count and not package_id_count_list[package_id]:
                    continue
                res[package_id][source][_id] = (format(float(count) / float(per_total_count), '.1%'), format(float(count) / float(package_id_count_list[package_id]), '.1%'))
            res[package_id][source] = sorted(res[package_id][source].items(), key=lambda x: float(x[1][0].replace('%', '')), reverse=True)
    except Exception as e:
        print(traceback.print_exc())
        print(package_id, source)
    return res


def to_csv(res):
    excel = openpyxl.Workbook()
    sheet = excel.create_sheet('1')
    count = 1
    sheet.cell(row=count, column=1, value='package_id')
    sheet.cell(row=count, column=2, value='source')
    sheet.cell(row=count, column=3, value='错误码')
    sheet.cell(row=count, column=4, value='比例')
    sheet.cell(row=count, column=5, value='影响比例')
    sheet.cell(row=count, column=6, value='错误码')
    sheet.cell(row=count, column=7, value='比例')
    sheet.cell(row=count, column=8, value='影响比例')
    sheet.cell(row=count, column=9, value='错误码')
    sheet.cell(row=count, column=10, value='比例')
    sheet.cell(row=count, column=11, value='影响比例')


    for package_id, content in res.items():
        for source, count_dict in content.items():
            count += 1
            sheet.cell(row=count, column=1, value=package_id)
            sheet.cell(row=count, column=2, value=source)
            column_count = 2
            for _id, value in count_dict:
                if _id is None:
                    continue
                if column_count == 11:
                    break
                column_count += 1
                print(column_count)
                sheet.cell(row=count, column=column_count, value=_id)
                column_count += 1
                if _id == 0 and float(value[0].replace('%', '')) < 95:
                    sheet.cell(row=count, column=column_count, value=value[0]).fill=styles.PatternFill(fill_type='solid',fgColor="00FF0000")
                elif _id != 0 and float(value[0].replace('%', '')) > 5:
                    sheet.cell(row=count, column=column_count, value=value[0]).fill=styles.PatternFill(fill_type='solid',fgColor="00FF0000")
                else:
                    sheet.cell(row=count, column=column_count, value=value[0])
                column_count += 1
                if _id != 0 and float(value[1].replace('%', '')) > 1:
                    sheet.cell(row=count, column=column_count, value=value[1]).fill=styles.PatternFill(fill_type='solid',fgColor="00FF0000")
                else:
                    sheet.cell(row=count, column=column_count, value=value[1])

    excel.save('error_source.xlsx')


error_code_by_source_by_rank()