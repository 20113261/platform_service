#coding:utf-8
import pymongo
import json
from count import TestCount
import sys
import pandas
import datetime
import math
import csv
import openpyxl



dept_city_list = ['北京',	'广州',	'上海',	'杭州',	'南京',	'苏州',	'天津',	'深圳',	'无锡',	'成都',	'重庆',	'武汉',	'沈阳',	'宁波',	'西安',	'郑州',	]
dest_city_list = ['东京',	'大阪',	'北海道',	'冲绳',	'福冈',	'京都',	'奈良',	'曼谷',	'芭堤雅',	'普吉岛',	'苏梅岛',	'清迈',	'新加坡',	'吉隆坡',	'清莱',	'华欣',	'沙巴',	'岘港',	'塞班',	'民丹岛',	'花莲',	'斯米兰',	'沙美岛',	'帕劳',	'巴厘岛',	'斐济',	'毛里求斯',	'塞舌尔',	'大溪地',	'芽庄',	'宿务',	'圣乔治',	'奥克兰',	'皇后镇',	'基督城',	'凯恩斯',	'悉尼',	'墨尔本',	'布里斯班',	'黄金海岸',	'圣保罗',	'乌斯怀亚',	'布宜诺斯艾利斯',	'里约热内卢',	'马瑙斯',	'加拉帕戈斯',	'墨西哥城',	'哈瓦那',	'伊斯坦布尔',	'迪拜',	'南安普顿',	'开罗',	'卡萨布兰卡',	'舍夫沙万',	'多伦多',	'蒙特利尔',	'渥太华',	'夏威夷',	'洛杉矶',	'拉斯维加斯',	'旧金山',	'奥兰多',	'芝加哥',	'休斯顿',	'华盛顿',	'纽约',	'波哥大',	'莫斯科',	'圣彼得堡',	'贝加尔湖',	'慕尼黑',	'法兰克福',	'柏林',	'科隆',	'杜塞尔多夫',	'巴黎',	'尼斯',	'阿维尼翁',	'摩纳哥',	'苏黎世',	'日内瓦',	'因特拉肯',	'罗马',	'威尼斯',	'佛罗伦萨',	'那不勒斯',	'布鲁塞尔',	'阿姆斯特丹',	'伦敦',	'爱丁堡',	'曼彻斯特',	'都柏林',	'赫尔辛基',	'奥斯陆',	'松恩峡湾',	'斯德哥尔摩',	'哥德堡',	'哥本哈根',	'雷克亚未克',	'巴塞罗那',	'马德里',	'里斯本',	'雅典',	'圣托里尼',	'伊斯坦布尔',	'维也纳',	'萨尔茨堡',	'因斯布鲁克',	'布达佩斯',	'布拉格',	'皮尔森',	'布拉迪斯拉发',	'华沙',	'卢森堡',	'杜布罗夫尼克',	'十六湖',	'卢布尔雅那',	'列支敦士登']


class MarketInfo(TestCount):
    def __init__(self, db, coll):
        super(MarketInfo, self).__init__(db, coll)

    def create_index(self):
        self.collections.create_index([('pid', 1)])

    def get_content(self):
        print_i = 0
        client = pymongo.MongoClient('mongodb://root:miaoji1109-=@10.19.2.103:27017/')
        collections = client['data_result']['tuniuGT_detail']
        for i in collections.find({}):
            print_i += 1
            print(print_i)
            result = i['result']
            if not result:
                continue
            result_0 = result[0]
            if not result_0:
                continue
            dept_city = result_0['dept']
            if not dept_city:
                continue
            dept_city = dept_city[0]['name']
            if dept_city:
                dept_city = dept_city.encode('utf-8')
            if dept_city not in dept_city_list:
                continue
            dest_city = result_0['dest']
            if not dest_city:
                continue
            dest_city = dest_city[0]['name']
            if dest_city:
                dest_city = dest_city.encode('utf-8')
            if dest_city not in dest_city_list:
                continue
            pid = result_0['pid_3rd']

            ptid = result_0['sid']  #携程是ptid，途牛是sid

            star_level = result_0['star_level']

            price_total = 0.00
            price_times = 0
            for r in result:
                price_list = r['tourist']
                for p in price_list:
                    price = p['price']  # if price else 0
                    if price:
                        price_times += 1
                        price_total += price
            price_avg = price_total / price_times
            data = {'pid': pid, 'dept_city': dept_city, 'dest_city': dest_city, 'price': price, 'ptid': ptid, 'star_level': star_level}
            self.tasks.append(data)
            self.pre_offset += 1

            self.insert_task()
        self.insert_mongo()

    def query(self):
        for line in self.collections.find({}):
            yield line

    def to_excel(self):
        data = []
        df_list = []
        columns = ['pid', 'dept_city', 'dest_city', 'price', 'ptid', 'star_level']
        #第一种
        # for i in self.query():
        #     res = {'pid': i['pid'], 'dept_city': i['dept_city'], 'dest_city': i['dest_city'], 'price':i['price'], 'ptid':i['ptid'], 'star_level':i['star_level']}
        #     data.append(res)
        #     if len(data) >= 2000:
        #         df = pandas.DataFrame(columns=columns,
        #                               data=data)
        #         df_list.append(df)
        #         data = []
        # pandas.concat(df_list).to_csv(datetime.datetime.now().strftime('./market_info%Y%m%d.csv'), encoding='utf-8')
        #第二种
        # with open('./test2.csv', 'w') as csv_file:
        #     writer = csv.DictWriter(csv_file, fieldnames=columns)
        #     # writer.writerows(dict(zip(columns, columns)))
        #     writer.writeheader()
        #     for i in self.query():
        #         # res = {i['pid'].encode('utf-8'), i['dept_city'].encode('utf-8'), i['dest_city'].encode('utf-8'), i['price'], i['ptid'].encode('utf-8'),i['star_level']]
        #         res = {'pid': i['pid'].encode('utf-8'), 'dept_city': i['dept_city'].encode('utf-8'), 'dest_city': i['dest_city'].encode('utf-8'), 'price':i['price'], 'ptid':i['ptid'].encode('utf-8'), 'star_level':i['star_level']}
        #
        #         data.append(res)
        #         if len(data) >= 2000:
        #             writer.writerows(data)
        #             data = []

        excel = openpyxl.Workbook()
        sheet = excel.create_sheet('1')
        count = 1
        sheet.cell(row=count, column=1, value='产品id')
        sheet.cell(row=count, column=2, value='出发城市')
        sheet.cell(row=count, column=3, value='目的城市')
        sheet.cell(row=count, column=4, value='产品日均价')
        sheet.cell(row=count, column=5, value='供应商名称')
        sheet.cell(row=count, column=6, value='产品等级')
        for i in self.query():
            count += 1
            print(count)
            # sheet.cell(row=count, column=1, value='4')

            sheet.cell(row=count, column=1, value=i['pid'])
            sheet.cell(row=count, column=2, value=i['dept_city'])
            sheet.cell(row=count, column=3, value=i['dest_city'])
            sheet.cell(row=count, column=4, value=i['price'])
            sheet.cell(row=count, column=5, value=i['ptid'])
            sheet.cell(row=count, column=6, value=i['star_level'])

        excel.save('market_info.xlsx')
        print(sheet.max_row)

if __name__ == '__main__':
    info = MarketInfo('data_result', 'tuniuGT_market_info')
    # info.get_content()
    info.to_excel()